"""Excel 报表导出：使用 openpyxl，无需 MS Office"""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:  # 没装时给出温和提示
    HAS_OPENPYXL = False


COLUMNS = [
    ("id",                    "序号",       8),
    ("datetime",              "检测时间",   20),
    ("sample_source",         "图像来源",   12),
    ("target_disease",        "病虫害类别", 16),
    ("confidence",            "置信度",     10),
    ("severity_level",        "严重程度",   10),
    ("count_value",           "数量",       8),
    ("phenophase",            "物候期",     16),
    ("farmer_name",           "农户姓名",   14),
    ("orchard_block",         "果园区块",   14),
    ("prescription_snapshot", "化学处方快照", 36),
    ("image_path",            "原图路径",   30),
    ("annotated_path",        "标注图路径", 30),
    ("notes",                 "备注",       18),
]


def export_logs(rows: Iterable[dict], out_path: Path) -> Path:
    """把若干检测日志导出为 .xlsx。"""
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "openpyxl 未安装，无法导出 Excel。请运行：pip install openpyxl"
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "廉江红橙检测台账"

    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    severity_fill = {
        "Green": PatternFill(start_color="DCEDC8", end_color="DCEDC8", fill_type="solid"),
        "Amber": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
        "Red":   PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    }

    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            v = r.get(key, "")
            if key == "confidence" and isinstance(v, (int, float)):
                v = f"{v * 100:.1f}%" if v <= 1.0 else f"{v:.1f}"
            cell = ws.cell(row=row_idx, column=col_idx, value=v if v is not None else "")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sev = r.get("severity_level")
        if sev in severity_fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = severity_fill[sev]

    ws.freeze_panes = "A2"
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path

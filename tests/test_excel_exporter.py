"""Excel 导出测试：跳过没装 openpyxl 的环境"""

import importlib

import pytest


openpyxl = importlib.util.find_spec("openpyxl")


@pytest.mark.skipif(openpyxl is None, reason="openpyxl 未安装")
def test_export_logs_basic(tmp_path):
    from lingnan.data.excel_exporter import export_logs
    rows = [
        {
            "id": 1, "datetime": "2026-05-28 10:00:00",
            "sample_source": "本地导入", "target_disease": "柑橘红蜘蛛",
            "confidence": 0.91, "severity_level": "Amber", "count_value": 8,
            "farmer_name": "张三", "orchard_block": "A-1", "phenophase": "挂果期",
            "prescription_snapshot": "[{\"name\":\"螺螨酯\"}]",
            "image_path": "/tmp/a.jpg", "annotated_path": "/tmp/a_det.jpg",
            "notes": "",
        },
        {
            "id": 2, "datetime": "2026-05-28 11:00:00",
            "sample_source": "本地导入", "target_disease": "柑橘黄龙病",
            "confidence": 0.78, "severity_level": "Red", "count_value": 1,
            "farmer_name": "李四", "orchard_block": "B-3", "phenophase": "采收期",
            "prescription_snapshot": "[]",
            "image_path": "/tmp/b.jpg", "annotated_path": "/tmp/b_det.jpg",
            "notes": "",
        },
    ]
    out = tmp_path / "test.xlsx"
    result = export_logs(rows, out)
    assert result.exists()
    assert result.stat().st_size > 0

    # 进一步校验内容
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    ws = wb.active
    # 表头 + 2 行
    rows_n = sum(1 for _ in ws.iter_rows(min_row=1))
    assert rows_n == 3
